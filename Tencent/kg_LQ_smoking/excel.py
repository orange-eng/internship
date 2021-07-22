
import openpyxl
#data = openpyxl.load_workbook('./inference/output/smoke_true_img123.xlsx')
# print(data.get_named_ranges()) # 输出工作页索引范围
# print(data.get_sheet_names()) # 输出所有工作页的名称
# # 取第一张表
# sheetnames = data.get_sheet_names()
# table = data.get_sheet_by_name(sheetnames[0])
# table = data.active
# print(table.title) # 输出表名
# #nrows = table.max_row # 获得行数
# nrows = 0
# ncolumns = table.max_column # 获得行数
# print(nrows,ncolumns)
# values = ['E','X','C','E','L']
# for value in values:
#     table.cell(nrows+1,6).value = value
#     nrows = nrows + 1
# data.save('./inference/output/smoke_true_img.xlsx')
import xlsxwriter
# Create a workbook and add a worksheet.
workbook = xlsxwriter.Workbook('Expenses02.xlsx')
worksheet = workbook.add_worksheet()
# # Some data we want to write to the worksheet.
# expenses = (
# ['Rent', 1000],
# ['Gas', 100],
# ['Food', 300],
# ['Gym', 50],
# )

# # Start from the first cell. Rows and columns are zero indexed.
# row = 0
# col = 0

# # Iterate over the data and write it out row by row.
# for item, cost in (expenses):
#     worksheet.write(row, col, item)
#     worksheet.write(row, col + 1, cost)
#     row += 1
#     # Write a total using a formula.
#     worksheet.write(row, 0, 'Total')
#     worksheet.write(row, 1, '=SUM(B1:B4)')
workbook.close()
