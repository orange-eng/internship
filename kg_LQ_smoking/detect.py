# Python
# author: Orange
# Phone: 17802005679
# Date: 2021.5.18

import argparse
from torch._C import Value

import torch.backends.cudnn as cudnn
from xlsxwriter import workbook

from models.experimental import *
from utils.datasets import *
from utils.utils import *
import dlib
import os
import sys
original_path = os.path.abspath(os.path.dirname(sys.argv[0]))
from imutils import face_utils
import openpyxl
import xlsxwriter       # create and write excel files


def detect(save_img=False,output_file=None,source_file=None):

    shape_predictor = original_path + "/face_shape/shape_predictor_68_face_landmarks.dat"
    detector = dlib.get_frontal_face_detector()
    predictor = dlib.shape_predictor(shape_predictor)
    out, source, weights, view_img, save_txt, imgsz = \
        opt.output, opt.source, opt.weights, opt.view_img, opt.save_txt, opt.img_size
    
    out = output_file
    source = source_file
    print('out=',out)
    print("source=",source)

    webcam = source == '0' or source.startswith('rtsp') or source.startswith('http') or source.endswith('.txt')

    # Initialize
    device = torch_utils.select_device(opt.device)
    if os.path.exists(out):
        shutil.rmtree(out)  # delete output folder
    os.makedirs(out)  # make new output folder
    half = device.type != 'cpu'  # half precision only supported on CUDA

    # Load model
    model = attempt_load(weights, map_location=device)  # load FP32 model
    imgsz = check_img_size(imgsz, s=model.stride.max())  # check img_size
    if half:
        model.half()  # to FP16

    # Second-stage classifier
    classify = False
    if classify:
        modelc = torch_utils.load_classifier(name='resnet101', n=2)  # initialize
        modelc.load_state_dict(torch.load('weights/resnet101.pt', map_location=device)['model'])  # load weights
        modelc.to(device).eval()

    # Set Dataloader
    vid_path, vid_writer = None, None
    if webcam:
        view_img = True
        cudnn.benchmark = True  # set True to speed up constant image size inference
        dataset = LoadStreams(source, img_size=imgsz)
    else:
        save_img = True
        dataset = LoadImages(source, img_size=imgsz)

    # Get names and colors
    names = model.module.names if hasattr(model, 'module') else model.names
    colors = [[random.randint(0, 255) for _ in range(3)] for _ in range(len(names))]

    # Run inference
    t0 = time.time()
    img = torch.zeros((1, 3, imgsz, imgsz), device=device)  # init img
    _ = model(img.half() if half else img) if device.type != 'cpu' else None  # run once
    
    
    
    detect_flag = 0         #记录被检测到的图片数目，如果识别到，则+1
    image_total = 0         #记录一个文件中的照片总数

    for path, img, im0s, vid_cap in dataset:
        img = torch.from_numpy(img).to(device)
        img = img.half() if half else img.float()  # uint8 to fp16/32
        img /= 255.0  # 0 - 255 to 0.0 - 1.0
        if img.ndimension() == 3:
            img = img.unsqueeze(0)

        # Inference
        t1 = torch_utils.time_synchronized()
        pred = model(img, augment=opt.augment)[0]
        #print("pred_shape:", pred.shape)
        #print("pred", pred[0][:2])
        # Apply NMS
        pred = non_max_suppression(pred, opt.conf_thres, opt.iou_thres, classes=opt.classes, agnostic=opt.agnostic_nms)
        t2 = torch_utils.time_synchronized()

        # Apply Classifier
        if classify:
            pred = apply_classifier(pred, modelc, img, im0s)

        
        # Process detections
        for i, det in enumerate(pred):  # detections per image
            image_total = image_total + 1
            if image_total % 1 == 0:
                if webcam:  # 电脑自带相机,batch_size >= 1
                    p, s, im0 = path[i], '%g: ' % i, im0s[i].copy()
                else:
                    p, s, im0 = path, '', im0s

                save_path = str(Path(out) / Path(p).name)
                print("save_path=",save_path)
                txt_path = str(Path(out) / Path(p).stem) + ('_%g' % dataset.frame if dataset.mode == 'video' else '')
                s += '%gx%g ' % img.shape[2:]  # print string
                gn = torch.tensor(im0.shape)[[1, 0, 1, 0]]  # normalization gain whwh
                if det is not None and len(det):
                    # Rescale boxes from img_size to im0 size
                    det[:, :4] = scale_coords(img.shape[2:], det[:, :4], im0.shape).round()

                    # Print results
                    # for c in det[:, -1].unique():
                    #     n = (det[:, -1] == c).sum()  # detections per class
                    #     s += '%g %ss, ' % (n, names[int(c)])  # add to string

                    # Write results
                    for *xyxy, conf, cls in det:    #烟头的坐标保存在xyxy里面
                        if save_txt:  # Write to file
                            xywh = (xyxy2xywh(torch.tensor(xyxy).view(1, 4)) / gn).view(-1).tolist()  # normalized xywh
                            with open(txt_path + '.txt', 'a') as f:
                                f.write(('%g ' * 5 + '\n') % (cls, *xywh))  # label format

                        mouth_y,mouth_h,mouth_x,mouth_w = 0,0,0,0
                        if save_img or view_img:  # Add bbox to image
                            label = '%s %.2f' % (names[int(cls)], conf)
                            
                            ############################
                            # 嘴唇与烟头的交集区域判断逻辑
                            ############################

                            rects = detector(im0,1)
                            for (i, rect) in enumerate(rects):
                                # 确定面部区域进行面部标志检测,并将其检测到的68个点转换为方便python处理的Numpy array
                                gray = cv2.cvtColor(im0, cv2.COLOR_BGR2GRAY)
                                shape = predictor(gray, rect)
                                shape = face_utils.shape_to_np(shape)

                            
                                # 循环遍历面部标志独立的每一部分
                                for (name, (i, j)) in face_utils.FACIAL_LANDMARKS_IDXS.items():
                                    if name == "inner_mouth":
                                        # 复制一张原始图的拷贝,以便于绘制面部区域,及其名称
                                        clone = im0.copy()
                                        # 遍历独立的面部标志的每一部分包含的点,并画在图中
                                        for (x, y) in shape[i:j]:
                                            #cv2.circle(im0, (x, y), 1, (0, 0, 255), -1)
                                            # 要实际提取每个面部区域，我们只需要计算与特定区域关联的（x，y）坐标的边界框，并使用NumPy数组切片来提取它：
                                            (x, y, w, h) = cv2.boundingRect(np.array([shape[i:j]]))
                                            #cv2.rectangle(im0,(x,y),(x+w,y+h),(0,0,255),thickness=1)
                                            mouth_y,mouth_h,mouth_x,mouth_w = y,h,x,w
                            
                            ########### 有嘴唇检测
                            #提取烟头方框的坐标数值,为左上角和右下角的坐标
                            smoke_left_up,smoke_right_down=(int(xyxy[0]), int(xyxy[1])), (int(xyxy[2]), int(xyxy[3]))
                            if xyxy[3] < mouth_y or xyxy[1] > mouth_y+mouth_h or xyxy[0]>mouth_x+mouth_w or xyxy[2]<mouth_x:
                                if conf > 0.85:  #在距离嘴巴比较远的位置时，如果conf>0.8则判断为烟头
                                    plot_one_box(xyxy, im0, label=label, color=colors[int(cls)], line_thickness=3)      
                                    detect_flag = detect_flag + 1
                                else:   
                                    print('')
                                    #print("NO drawing. It is not smoking!.......................................")
                            else:
                                if conf >0.75:   #在距离嘴巴比较近的位置时，如果conf>0.7，才判断为烟头   
                                    #print("drawing this retrangle................................")
                                    detect_flag = detect_flag + 1
                                    plot_one_box(xyxy, im0, label=label, color=colors[int(cls)], line_thickness=3)
                            
                            ########### 无嘴唇检测
                            # if conf > 0.61:  #在距离嘴巴比较远的位置时，如果conf>0.8则判断为烟头
                            #     # plot_one_box(xyxy, im0, label=label, color=colors[int(cls)], line_thickness=3)      
                            #     detect_flag = detect_flag + 1

                            ##########################################
                            ### 结束
                            ##########################################

                # Save results (image with detections)
                if save_img:
                    if dataset.mode == 'images':
                        cv2.imwrite(save_path, im0)
                    else:
                        if vid_path != save_path:  # new video
                            vid_path = save_path
                            if isinstance(vid_writer, cv2.VideoWriter):
                                vid_writer.release()  # release previous video writer

                            fourcc = 'mp4v'  # output video codec
                            fps = vid_cap.get(cv2.CAP_PROP_FPS)
                            w = int(vid_cap.get(cv2.CAP_PROP_FRAME_WIDTH))
                            h = int(vid_cap.get(cv2.CAP_PROP_FRAME_HEIGHT))
                            vid_writer = cv2.VideoWriter(save_path, cv2.VideoWriter_fourcc(*fourcc), fps, (w, h))
                        vid_writer.write(im0)

    # --------------------------- 打印每一个视频的数据    
    if save_txt or save_img:
        print('Results saved to %s' % os.getcwd() + os.sep + out)
        if platform == 'darwin' and not opt.update:  # MacOS
            os.system('open ' + save_path)
    print('Done. (%.3fs)' % (time.time() - t0))

    print("detect_flag=",detect_flag)
    print("images_total = ", image_total)

    ############ 有嘴唇检测
    if detect_flag >= 2 and detect_flag<45:
        print("This video is smoking!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!")
        return True,detect_flag,image_total
    else:
        print("This video is no smoking.")
        return False,detect_flag,image_total
    
    ########### 无嘴唇检测
    # if detect_flag > 0:
    #     print("This video is smoking!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!")
    #     return True,detect_flag,image_total
    # else:
    #     print("This video is no smoking.")
    #     return False,detect_flag,image_total


class Detect:
    def __init__(self, opt):

        weights, imgsz = opt.weights, opt.img_size

        # Initialize
        self.device = torch_utils.select_device(opt.device)

        self.ishalf = self.device.type != 'cpu'  # half precision only supported on CUDA

        # Load model
        self.model = attempt_load(weights, map_location=self.device)  # load FP32 model
        self.imgsz = check_img_size(imgsz, s=self.model.stride.max())  # check img_size
        if self.ishalf:
            self.model.half()  # to FP16

        self.names = self.model.module.names if hasattr(self.model, 'module') else self.model.names
        # colors = [[random.randint(0, 255) for _ in range(3)] for _ in range(len(names))]

        # source = cv2.imread('./inference/images/bus.jpg')


    def detect_v2(self, source):
        img, im0s = LoadImages_v2(source, img_size=self.imgsz)

        # Run inference
        t0 = time.time()
        img_ten = torch.zeros((1, 3, self.imgsz, self.imgsz), device=self.device)  # init img
        _ = self.model(img_ten.half() if self.ishalf else img_ten) if self.device.type != 'cpu' else None  # run once

        img = torch.from_numpy(img).to(self.device)
        img = img.half() if self.ishalf else img.float()  # uint8 to fp16/32
        img /= 255.0  # 0 - 255 to 0.0 - 1.0
        if img.ndimension() == 3:
            img = img.unsqueeze(0)

        # Inference
        t1 = torch_utils.time_synchronized()
        pred = self.model(img, augment=opt.augment)[0]
        print("pred", pred.shape)


        # Apply NMS
        pred = non_max_suppression(pred, opt.conf_thres, opt.iou_thres, classes=[0,27], agnostic=opt.agnostic_nms)
        t2 = torch_utils.time_synchronized()

        # Process detections
        for i, det in enumerate(pred):  # detections per image

            # print("i",i)
            s, im0 = '', im0s

            s += '%gx%g ' % img.shape[2:]  # print string
            # gn = torch.tensor(im0.shape)[[1, 0, 1, 0]]  # normalization gain whwh
            if det is not None and len(det):
                # Rescale boxes from img_size to im0 size
                det[:, :4] = scale_coords(img.shape[2:], det[:, :4], im0.shape).round()
                # print(det)

                for *bbox, conf, cls in det:
                    # bbox = [int(i) for i in xyxy]
                    # print(bbox)
                    crop_img = im0[int(bbox[1]):int(bbox[3]),int(bbox[0]):int(bbox[2])]
                    a = str(random.randint(0,100))
                    cv2.imwrite(a+"crop_img.jpg",crop_img)
                    # label = '%s %.2f' % (self.names[int(cls)], conf)
                    # print(label, a)
            print('Done. (%.3fs)' % (time.time() - t0))


if __name__ == '__main__':

    # images_path = "inference\images\\50_50_test\\False_image"        # 存放照片文件的位置
    images_path = "inference\images\\6"        # 存放照片文件的位置
    
    images_file =  os.listdir(images_path)  # 读取每一个文件夹下面的文件名
    # print(images_file)

    
    for j in range(len(images_file)):
        parser = argparse.ArgumentParser()
        folder = 1
        parser.add_argument('--weights', nargs='+', type=str, default=r'runs\evolution\weights\best.pt', help='model.pt path(s)')
        # parser.add_argument('--source', type=str, default='./inference/images/smoke_false_img/{}'.format(folder), help='source')  # file/folder, 0 for webcam
        # parser.add_argument('--output', type=str, default='inference/output/smoke_false_img/{}'.format(folder), help='output folder')  # output folder

        # parser.add_argument('--source', type=str, default='./inference/images/50_50_test/False_image/{}/'.format(images_file[j]), help='source')  # file/folder, 0 for webcam
        # parser.add_argument('--output', type=str, default='inference/output/50_50_test/False_image/{}/'.format(images_file[j]), help='output folder')  # output folder
        # parser.add_argument('--excel_file', type=str, default='./inference/output/50_50_test/False_image/{}.xlsx'.format(images_file[j]), help='excel folder')  # output folder
        parser.add_argument('--source', type=str, default='./inference/images/6/{}/'.format(images_file[j]), help='source')  # file/folder, 0 for webcam
        parser.add_argument('--output', type=str, default='inference/output/6/{}/'.format(images_file[j]), help='output folder')  # output folder
        parser.add_argument('--excel_file', type=str, default='./inference/output/6/{}.xlsx'.format(images_file[j]), help='excel folder')  # output folder

        parser.add_argument('--img-size', type=int, default=640, help='inference size (pixels)')
        parser.add_argument('--conf-thres', type=float, default=0.36, help='object confidence threshold')
        parser.add_argument('--iou-thres', type=float, default=0.4, help='IOU threshold for NMS')
        parser.add_argument('--device', default='0', help='cuda device, i.e. 0 or 0,1,2,3 or cpu')
        parser.add_argument('--view-img', action='store_true', help='display results')
        parser.add_argument('--save-txt', action='store_true', help='save results to *.txt')
        parser.add_argument('--classes', type=int, help='filter by class: --class 0, or --class 0 2 3')
        parser.add_argument('--agnostic-nms', action='store_true', help='class-agnostic NMS')
        parser.add_argument('--augment', action='store_true', default=True, help='augmented inference')
        parser.add_argument('--update', action='store_true', help='update all models')
        opt = parser.parse_args()

        #create excel files 
        # workbook = xlsxwriter.Workbook(opt.excel_file)
        # worksheet = workbook.add_worksheet()
        # workbook.close()
        
        #-----------------------
        # print("Writing excel files {}".format(opt.excel_file))

        with torch.no_grad():
            if opt.update:  # update all models (to fix SourceChangeWarning)
                for opt.weights in ['yolov5s.pt', 'yolov5m.pt', 'yolov5l.pt', 'yolov5x.pt', 'yolov3-spp.pt']:
                    detect()
                    create_pretrained(opt.weights, opt.weights)
            else:

                # detector = Detect(opt)
                # path = './inference/images/'
                # for i, file in enumerate(os.listdir(path)):
                #     img = cv2.imread(path + file)
                #     detector.detect_v2(img)
                # opt.classes = [0]

                all_files = os.listdir(opt.source)
                # print((all_files))

                
                total_num = len(all_files)
                # print(total_num)
                
                Smoke_video_num = 0     # 被检测是抽烟视频的数目
                picture_num = 0         # 被检测的照片数目
                picture_total = 0       # total amount of pictures in one file
                # data = openpyxl.load_workbook(opt.excel_file)
                # for i in range(total_num):
                # source_file = opt.source + "{}".format(all_files[i])
                # output_file = opt.output + "{}".format(all_files[i])

                start = time.time()
                Smoke_flag,picture_num,picture_total = detect(output_file=opt.output,source_file=opt.source)
                end = time.time()
                print("time consume:",end - start)
                
                ####################
                # 写入excel文件
                ####################

                # 取第一张表
                # sheetnames = data.get_sheet_names()
                # table = data.get_sheet_by_name(sheetnames[0])
                # table = data.active
                # value = 1
                # table.cell(i+1,1).value = all_files[i]  # 视频名
                # table.cell(i+1,2).value = Smoke_flag    # 检测结果
                # table.cell(i+1,3).value = "False"
                # table.cell(i+1,4).value = picture_num   # 识别为抽烟的帧率
                # table.cell(i+1,5).value = picture_total
                # data.save(opt.excel_file)
                # 输出检测结果
                if Smoke_flag:
                    Smoke_video_num = Smoke_video_num + 1
                    #print("Smoke_video_num = ",Smoke_video_num)
            print("Smoke_video_num = ",Smoke_video_num)
                
        # from queue import Queue
        # from threading import Thread
        # import time
        #
        # que = Queue(maxsize=10)
        #
        # def post(num):
        #     print(num)
        #
        #
        # while True:
        #     num = random.randint(1,100)
        #
        #     que.put(num)
        #
        #     threads = []
        #     if que.full():
        #         print(que.qsize())
        #         tt = time.time()
        #         for i in range(10):
        #             t = Thread(target=post, args=(que.get(),))  # post發送數據到後端
        #             t.start()
        #             t.join()
        #             # post(que.get())
        #         print(time.time()-tt)
        #     print("que.qsize()",que.qsize())

